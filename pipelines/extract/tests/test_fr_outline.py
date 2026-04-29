"""Tests for pipelines/extract/fr_outline.py"""

import datetime
from pathlib import Path

import duckdb
import openpyxl
import pytest

from pipelines.extract.fr_outline import (
    normalize_commune_name,
    parse_xlsx,
)


# ---------------------------------------------------------------------------
# Unit tests: normalize_commune_name
# ---------------------------------------------------------------------------

def test_normalize_basic():
    assert normalize_commune_name("Saint-Malo") == "SAINT MALO"


def test_normalize_accents():
    assert normalize_commune_name("Évreux") == "EVREUX"


def test_normalize_trailing_la():
    assert normalize_commune_name("Chapelle (LA)") == "CHAPELLE"


def test_normalize_trailing_les():
    assert normalize_commune_name("Sables (LES)") == "SABLES"


def test_normalize_trailing_le():
    assert normalize_commune_name("Havre (LE)") == "HAVRE"


def test_normalize_apostrophe():
    assert normalize_commune_name("L'Isle-Adam") == "L ISLE ADAM"


def test_normalize_collapse_spaces():
    assert normalize_commune_name("  Saint  Pierre  ") == "SAINT PIERRE"


def test_normalize_none():
    assert normalize_commune_name(None) == ""


def test_normalize_empty():
    assert normalize_commune_name("") == ""


# ---------------------------------------------------------------------------
# Synthetic xlsx helpers
# ---------------------------------------------------------------------------

def _make_bretagne_xlsx(tmp_path: Path) -> Path:
    """Synthetic Bretagne-like file (dept col, CVM col with decimal-comma + sentinels)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CVM"
    ws.append(["Département", "Commune", "Date", "CVM (µg/L)"])
    ws.append(["029", "Brest", datetime.date(2023, 6, 1), 0.25])
    ws.append(["029", "Quimper", datetime.date(2023, 6, 2), "0,10"])
    ws.append(["029", "Landerneau", datetime.date(2023, 6, 3), "#EMPTY"])
    ws.append(["029", "Morlaix", datetime.date(2023, 6, 4), "<0.5"])
    ws.append(["029", "Chapelle (LA)", datetime.date(2023, 6, 5), 0.60])
    ws.append(["2A", "Ajaccio", datetime.date(2023, 6, 6), 0.05])
    out = tmp_path / "Annexe C (Bretagne).xlsx"
    wb.save(out)
    return out


def _make_normandie_xlsx(tmp_path: Path) -> Path:
    """Synthetic Normandie-like file (numeric dept, header wording matches real file)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Données"
    ws.append(["Dépt - Code", "PSV - Commune - Nom", "PLV - Date",
               "Chlorure de vinyl monomère (µg/L)"])
    ws.append([14, "Caen", datetime.date(2022, 1, 15), "0,55"])
    ws.append([76, "Rouen", datetime.date(2022, 3, 20), "0"])
    out = tmp_path / "Annexe F (Normandie).xlsx"
    wb.save(out)
    return out


def _make_aquitaine_xlsx(tmp_path: Path) -> Path:
    """Synthetic Nouvelle-Aquitaine-like file ('Résultat' column)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CVM"
    ws.append(["Dept", "Commune", "PLV - Date", "Résultat"])
    ws.append(["33", "Bordeaux", datetime.date(2021, 5, 10), "1,20"])
    ws.append(["33", "Pessac", datetime.date(2021, 5, 11), "#EMPTY"])
    out = tmp_path / "Annexe G (Nouvelle-Aquitaine).xlsx"
    wb.save(out)
    return out


@pytest.fixture()
def conn():
    c = duckdb.connect()
    yield c
    c.close()


# ---------------------------------------------------------------------------
# Integration tests: parse_xlsx
# ---------------------------------------------------------------------------

def test_parse_bretagne_basic(conn, tmp_path):
    path = _make_bretagne_xlsx(tmp_path)
    rows = parse_xlsx(conn, path)
    assert len(rows) == 6

    brest = next(r for r in rows if r["commune_name_raw"] == "Brest")
    assert brest["dept"] == "29"
    assert brest["value_ugl"] == pytest.approx(0.25)
    assert brest["commune_name_norm"] == "BREST"

    # Decimal-comma value parsed
    quimper = next(r for r in rows if r["commune_name_raw"] == "Quimper")
    assert quimper["value_ugl"] == pytest.approx(0.10)

    # Sentinels → None
    assert next(r for r in rows if r["commune_name_raw"] == "Landerneau")["value_ugl"] is None
    assert next(r for r in rows if r["commune_name_raw"] == "Morlaix")["value_ugl"] is None

    # (LA) stripped from norm
    chapelle = next(r for r in rows if r["commune_name_raw"] == "Chapelle (LA)")
    assert chapelle["commune_name_norm"] == "CHAPELLE"
    assert chapelle["value_ugl"] == pytest.approx(0.60)

    # Corsican dept preserved
    ajaccio = next(r for r in rows if r["commune_name_raw"] == "Ajaccio")
    assert ajaccio["dept"] == "2A"


def test_parse_normandie_numeric_dept(conn, tmp_path):
    path = _make_normandie_xlsx(tmp_path)
    rows = parse_xlsx(conn, path)
    assert len(rows) == 2
    caen = next(r for r in rows if r["commune_name_raw"] == "Caen")
    assert caen["dept"] == "14"
    assert caen["value_ugl"] == pytest.approx(0.55)
    rouen = next(r for r in rows if r["commune_name_raw"] == "Rouen")
    assert rouen["dept"] == "76"
    assert rouen["value_ugl"] == pytest.approx(0.0)


def test_parse_aquitaine_empty_sentinel(conn, tmp_path):
    path = _make_aquitaine_xlsx(tmp_path)
    rows = parse_xlsx(conn, path)
    assert len(rows) == 2
    bx = next(r for r in rows if r["commune_name_raw"] == "Bordeaux")
    assert bx["dept"] == "33"
    assert bx["commune_name_norm"] == "BORDEAUX"
    assert bx["value_ugl"] == pytest.approx(1.20)
    pessac = next(r for r in rows if r["commune_name_raw"] == "Pessac")
    assert pessac["value_ugl"] is None


def test_source_file_set(conn, tmp_path):
    path = _make_bretagne_xlsx(tmp_path)
    rows = parse_xlsx(conn, path)
    assert all(r["source_file"] == path.name for r in rows)


def test_plv_date_is_python_date(conn, tmp_path):
    path = _make_bretagne_xlsx(tmp_path)
    rows = parse_xlsx(conn, path)
    brest = next(r for r in rows if r["commune_name_raw"] == "Brest")
    assert isinstance(brest["plv_date"], datetime.date)


def test_missing_required_columns_returns_empty(conn, tmp_path):
    """File without dept/commune/date/value columns → no rows (skipped)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["foo", "bar"])
    ws.append([1, 2])
    p = tmp_path / "weird.xlsx"
    wb.save(p)
    assert parse_xlsx(conn, p) == []


def test_non_numeric_dept_skipped(conn, tmp_path):
    """File whose dept column holds department names → rows dropped."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Departement", "Commune", "Date", "CVM"])
    ws.append(["AIN", "Anglefort", datetime.date(2020, 1, 1), 0.1])
    ws.append(["ALLIER", "Moulins", datetime.date(2020, 2, 1), 0.2])
    p = tmp_path / "names_as_dept.xlsx"
    wb.save(p)
    assert parse_xlsx(conn, p) == []
